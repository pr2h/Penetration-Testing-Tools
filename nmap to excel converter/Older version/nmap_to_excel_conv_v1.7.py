# Welcome function
def hujambo():
    # Release Date : 21 May 2018
    print '''
##################################################
# Tool    : nmap_to_excel_conv                   #
# Version : 1.7                                  #
# Profile : https://github.com/pr2h/             #
# Coded with Python 2.7                          #
# ######          #####   #                      #
# #    #  # ###  #     #  #                      #
# ######  ##         #    ######                 #
# #       #        #      #    #                 #
# #       #      #######  #    #                 #
##################################################
    '''

# Required imports
import sys
import os.path

# Importing openpyxl for excel operations
try:
    from openpyxl import Workbook
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, PatternFill, Font
    from openpyxl.styles.borders import Border, Side

# ERROR if openpyxl is not installed
except:
    print "You don't have openpyxl, please perform 'pip install openpyxl'"
    print "or download and install it from 'https://pypi.python.org/pypi/openpyxl'"
    sys.exit()

def nmap_to_xl_exec(filename):
    # Initializing / Assigning variables
    count=2
    column_width_b = 0
    column_width_c = 0
    space_before_service = 10
    cell_border = Border(top=Side(style='medium'), bottom=Side(style='medium'), left=Side(style='medium'), right=Side(style='medium'))
    fileexists = 0

    while True:
        if os.path.isfile(filename.replace('.txt','')+'.xlsx') == True:
            fileexists = 1
            
            wb = load_workbook(filename.replace('.txt','')+'.xlsx')
            ws = wb.active
            first_column = ws['A']
            count = len(first_column)+1
            
                    
        else:
            fileexists = 0
            # Creating / Opening excel
            wb=Workbook()
            ws=wb.active
            ws.title="Open_Ports"

        break
          
    #except:
    #print "ERROR: Excel file is open / inaccessible.\nPlease close it and try again. Exiting.."
    #sys.exit()

    if fileexists == 0:
        # Initializing / Assigning Heading cells
        try:
            heading_cells = ['A1','B1','C1']
            ws[heading_cells[0]]='S.No.'
            ws[heading_cells[1]]='IP'
            ws[heading_cells[2]]='Open Ports'

            for heading_cell in heading_cells:
                ws[heading_cell].alignment = Alignment(horizontal='center', vertical='center')
                ws[heading_cell].fill = PatternFill(start_color='00C0C0C0', end_color='00C0C0C0', fill_type='solid')
                ws[heading_cell].border = cell_border
                ws[heading_cell].font = ws[heading_cell].font.copy(bold=True)
        except:
            print "ERROR: Excel file is open / inaccessible.\nPlease close it and try again. Exiting.."
            sys.exit()

    # Reading line by line
    try:
        for line in content:
            # Checking for new IP
            if line.startswith("Nmap scan report for "):
                cell="A"+str(count)
                ws[cell].value=count-1
                ws[cell].alignment = Alignment(horizontal='center', vertical='top')
                ws[cell].border = cell_border
                
                IP=line.replace("Nmap scan report for ",'')
                cell="B"+str(count)
                count+=1
                IP = IP.strip('\n')
                ws[cell].value=IP
                ws[cell].alignment = Alignment(horizontal='center', vertical='top')
                ws[cell].border = cell_border
                
                if len(IP) > column_width_b:
                    column_width_b = len(IP)

            # Checking for open ports
            elif "open" in line and "open|filtered" not in line:
                complete_line = ''
                line = line.strip('\n')
                line = line.replace('open','')
                line = line.replace('/tcp','/TCP')
                line = line.replace('/udp','/UDP')
                line = line.replace('unknown','-')
                line = line.split()
                line[0] = line[0]+((space_before_service-len(line[0]))*' ')
                complete_line = (line[0]+'\t'+line[1]).expandtabs(space_before_service)
                
                cell="C"+str(count-1)
                if ws[cell].value==None:
                    ws[cell].value='="'+complete_line+'"'
                    ws.row_dimensions[count-1].height=14.4
                else:
                    temp=ws[cell].value+'&CHAR(10)&"'+complete_line+'"'
                    ws[cell].value=temp
                    ws[cell].alignment = Alignment(wrapText=True)
                    ws.row_dimensions[count-1].height+=14.4

                ws[cell].border = cell_border
                
                if len(complete_line) > column_width_c:
                    column_width_c = len(complete_line)       

        # Assigning column dimensions
        ws.column_dimensions['A'].width = 6
        ws.column_dimensions['B'].width = column_width_b
        ws.column_dimensions['C'].width = column_width_c

        # Saving output file
        filename=filename.replace('.txt','')+'.xlsx'
        wb.save(filename)
        print 'File saved as : ',filename
        
    except:
        print "ERROR: Excel file is open / inaccessible.\nPlease close it and try again. Exiting.."
        sys.exit()

if __name__=='__main__':
    hujambo() # Welcome message

    # Reading nmap output (.txt file)
    while True:
        filename=raw_input("Enter filename: ")
            
        if filename.endswith('.txt'):
            pass
        else:
            filename=filename+'.txt'
                
        try:
            with open(filename,"r") as f:
                content=f.readlines()
            f.close()
            break
        except:
            print "No such file exists. Please enter filename (.txt) again"
            
    nmap_to_xl_exec(filename)
